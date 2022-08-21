import time
import unittest

from ddt import data, unpack, ddt
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.common.by import By


@ddt
class MyTestCase(unittest.TestCase):
    def setUp(self) -> None:
        self.driver = webdriver.Chrome("C:\\Selenium\\chromedriver.exe")
        driver = self.driver
        driver.maximize_window()
        driver.get("https://demoblaze.com")
    def read_data_from_excel(file_name,sheet):
        datalist = []
        wb = load_workbook(filename=file_name)
        sh = wb[sheet]
        row_ct = sh.max_row
        col_ct = sh.max_column

        for i in range(2, row_ct + 1):
            row = []
            for j in range(1, col_ct+1):
                row.append(sh.cell(row=1,column=1).value)
            datalist.append(row)
            return datalist

    @data(*read_data_from_excel("jetbrains://pycharm/navigate/reference?project=eveningautomation&path=testdata/Username_data.xlsx","Sheet1"))
    @unpack
    def test_login(self,Username,Password):
        driver = self.driver
        LogIn = driver.find_element(By.ID, "login2")
        driver.implicitly_wait(10)
        LogIn.click()
        loginUsername = driver.find_element(By.ID, "loginusername")
        loginUsername.send_keys(Username)
        loginPassword = driver.find_element(By.ID, "loginpassword")
        loginPassword.send_keys(Password)
        Button = driver.find_element(By.XPATH, '//*[@id="logInModal"]/div/div/div[3]/button[2]')
        Button.click()
        time.sleep(10)
        result = "Welcome testmorning"
        value = driver.find_element(By.ID, "nameofuser").text
        print(value)
        self.assertEqual(result, value, "Does not match so test case failed")



    def tearDown(self) -> None:
        self.driver.quit()


if __name__ == '__main__':
    unittest.main()
